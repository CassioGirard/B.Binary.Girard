from iqoptionapi.stable_api import IQ_Option
from sqlite import bd_sqlite
from noticias import NoticiasInvesting
import time
from configobj import ConfigObj
import json, sys
from datetime import datetime, timedelta
from catalogador import catag
from tabulate import tabulate
import pandas as pd
import openpyxl
#global valor_entrada
#global contador_loss_consecutivas 

### CRIANDO ARQUIVO DE CONFIGURAÇÃO ####
config = ConfigObj('config.txt')
email = config['LOGIN']['email']
senha = config['LOGIN']['senha']
tipo = config['AJUSTES']['tipo']
otc  = config['AJUSTES']['OTC']
valor_entrada = float(config['AJUSTES']['valor_entrada'])
stop_win = float(config['AJUSTES']['stop_win'])
stop_loss = float(config['AJUSTES']['stop_loss'])
lucro_total = 0
stop = True
contador_loss_consecutivas = 0
contador_loss_consecutivas_historico = 0
contador_win_consecutivas = 0
contador_win_consecutivas_historico = 0
# URL da página de noticias
url = "https://br.investing.com/economic-calendar/"
# Criar uma instância da classe NoticiasInvesting
noticias_investing = NoticiasInvesting()


if config['MARTINGALE']['usar_martingale'].upper() == 'S':
    martingale = int(config['MARTINGALE']['niveis_martingale'])
else:
    martingale = 0
fator_mg = float(config['MARTINGALE']['fator_martingale'])

if config['SOROS']['usar_soros'].upper() == 'S':
    soros = True
    niveis_soros = int(config['SOROS']['niveis_soros'])
    nivel_soros = 0

else:
    soros = False
    niveis_soros = 0
    nivel_soros = 0

valor_soros = 0
lucro_op_atual = 0

analise_medias = config['AJUSTES']['analise_medias']
velas_medias = int(config['AJUSTES']['velas_medias'])

print('Iniciando Conexão com a IQOption')
API = IQ_Option(email,senha)

### Função para conectar na IQOPTION ###
check, reason = API.connect()
if check:
    print('\nConectado com sucesso')
else:
    if reason == '{"code":"invalid_credentials","message":"You entered the wrong credentials. Please ensure that your login/password is correct."}':
        print('\nEmail ou senha incorreta')
        sys.exit()
        
    else:
        print('\nHouve um problema na conexão')

        print(reason)
        sys.exit()

### Função para Selecionar demo ou real ###
while True:
    #escolha = input('\nSelecione a conta em que deseja conectar: demo ou real  - ')
    escolha = 'demo'
    if escolha == 'demo':
        conta = 'PRACTICE'
        print('Conta demo selecionada')
        break
    if escolha == 'real':
        conta = 'REAL'
        print('Conta real selecionada')
        break
    else:
        print('Escolha incorreta! Digite demo ou real')
        
API.change_balance(conta)

### Função para checar stop win e loss
def check_stop(entrada):
    global stop,lucro_total,conta

    if lucro_total <= float('-'+str(abs(stop_loss))): # Se stop loss

        arquivo_excel = "resultado.xlsx"
        # Abrir ou criar o arquivo Excel
        try:
            wb = openpyxl.load_workbook(arquivo_excel)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        # Selecionar a primeira planilha
        ws = wb.active
        # Verificar se o cabeçalho já existe
        if ws.cell(row=1, column=1).value is None:
            ws.cell(row=1, column=1).value = "Win"
            ws.cell(row=1, column=2).value = "Loss"
            ws.cell(row=1, column=3).value = "Win TR"
            ws.cell(row=1, column=4).value = "Loss TR"
        if conta == 'PRACTICE':
            ws.cell(row=2, column=4).value = (ws.cell(row=2, column=4).value or 0) + 1
        else:
            ws.cell(row=2, column=2).value = (ws.cell(row=2, column=2).value or 0) + 1
            stop = False
            print('\n#########################')
            print('STOP LOSS BATIDO ',str(cifrao),str(lucro_total))
            print(f'\n>>>>>> Perdas consecutivas: {contador_loss_consecutivas_historico}')
            print('#########################')
            sys.exit() #ativar pós teste
        wb.save(arquivo_excel)
        

    if lucro_total >= float(abs(stop_win)): #STOP WIN

            # Definindo o nome do arquivo Excel
            arquivo_excel = "resultado.xlsx"
            # Abrir ou criar o arquivo Excel
            try:
                wb = openpyxl.load_workbook(arquivo_excel)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            # Selecionar a primeira planilha
            ws = wb.active
            # Verificar se o cabeçalho já existe
            if ws.cell(row=1, column=1).value is None:
                ws.cell(row=1, column=1).value = "Win"
                ws.cell(row=1, column=2).value = "Loss"
                ws.cell(row=1, column=3).value = "Win TR"
                ws.cell(row=1, column=4).value = "Loss TR"
            if conta == 'PRACTICE':
                ws.cell(row=2, column=3).value = (ws.cell(row=2, column=3).value or 0) + 1
                if contador_win_consecutivas_historico >= 3 and contador_loss_consecutivas_historico <=2:
                    print('trocado para a REAL')
                    conta='REAL'
                    API.change_balance(conta)
            else:
                ws.cell(row=2, column=1).value = (ws.cell(row=2, column=1).value or 0) + 1
                stop = False
                print('\n#########################')
                print('STOP WIN BATIDO ',str(cifrao),str(lucro_total))
                print(f'\n>>>>>> Perdas consecutivas: {contador_loss_consecutivas_historico}')
                print('#########################')
                sys.exit() #ativar pós teste
            wb.save(arquivo_excel)


    if entrada > float(stop_loss):
        # Definindo o nome do arquivo Excel
        arquivo_excel = "resultado.xlsx"
        # Abrir ou criar o arquivo Excel
        try:
            wb = openpyxl.load_workbook(arquivo_excel)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        # Selecionar a primeira planilha
        ws = wb.active
        # Verificar se o cabeçalho já existe
        if ws.cell(row=1, column=1).value is None:
            ws.cell(row=1, column=1).value = "Win"
            ws.cell(row=1, column=2).value = "Loss"
            ws.cell(row=1, column=3).value = "Win TR"
            ws.cell(row=1, column=4).value = "Loss TR"
        if conta == 'PRACTICE':
                ws.cell(row=2, column=4).value = (ws.cell(row=2, column=4).value or 0) + 1
        else:
            ws.cell(row=2, column=2).value = (ws.cell(row=2, column=2).value or 0) + 1
            print('\n#########################')
            print('A BANCA QUEBROU! ',str(cifrao),str(lucro_total))
            print(f'\n>>>>>> Perdas consecutivas: {contador_loss_consecutivas_historico}')
            print('#########################')
            sys.exit() #ativar pós teste
        wb.save(arquivo_excel)


def payout(par):
    profit = API.get_all_profit()
    all_asset = API.get_all_open_time()

    try:
        if all_asset['binary'][par]['open']:
            if profit[par]['binary']> 0:
                binary = round(profit[par]['binary'],2) * 100
        else:
            binary  = 0
    except:
        binary = 0

    try:
        if all_asset['turbo'][par]['open']:
            if profit[par]['turbo']> 0:
                turbo = round(profit[par]['turbo'],2) * 100
        else:
            turbo  = 0
    except:
        turbo = 0

    try:
        if all_asset['digital'][par]['open']:
            digital = API.get_digital_payout(par)
        else:
            digital  = 0
    except:
        digital = 0

    return binary, turbo, digital

### Função abrir ordem e checar resultado ###
def compra(ativo,valor_entrada, direcao, exp, tipo, i, resultado_win, resultado_loss, quantidade_compra):
    global stop,lucro_total, nivel_soros, niveis_soros, valor_soros, lucro_op_atual,contador_loss_consecutivas,contador_loss_consecutivas_historico,contador_win_consecutivas,contador_win_consecutivas_historico

    if soros:
        if nivel_soros == 0:
            entrada = valor_entrada

        if nivel_soros >=1 and valor_soros > 0 and nivel_soros <= niveis_soros:
            entrada = valor_entrada + valor_soros

        if nivel_soros > niveis_soros:
            lucro_op_atual = 0
            valor_soros = 0
            entrada = valor_entrada
            nivel_soros = 0
    else:
        entrada = valor_entrada

    for i in range(martingale + 1):

        if stop == True:
        
            if tipo == 'digital':
                check, id = API.buy_digital_spot_v2(ativo,entrada,direcao,exp)
            else:
                check, id = API.buy(entrada,ativo,direcao,exp)

            if check:
                if i == 0: 
                    print('\n>> Ordem aberta \n>> Par:',ativo,'\n>> Timeframe:',exp,'\n>> Entrada de:',cifrao,entrada)
                    quantidade_compra = quantidade_compra + 1
                if i >= 1:
                    print('\n>> Ordem aberta para gale',str(i),'\n>> Par:',ativo,'\n>> Timeframe:',exp,'\n>> Entrada de:',cifrao,entrada)
                    quantidade_compra = quantidade_compra + 1

                while True:
                    time.sleep(0.1)
                    status , resultado = API.check_win_digital_v2(id) if tipo == 'digital' else API.check_win_v4(id)

                    if status:

                        lucro_total += round(resultado,2)
                        valor_soros += round(resultado,2)
                        lucro_op_atual += round(resultado,2)
                        valorconta = float(API.get_balance())

                        if resultado > 0:
                            if i == 0:
                                print('\n>> Resultado: WIN \n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2), '\n>> Saldo conta:',cifrao,valorconta)

                            if i >= 1:
                                print('\n>> Resultado: WIN no gale',str(i),'\n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2),'\n>> Saldo conta:',cifrao,valorconta)
                            resultado_win = resultado_win + 1
                            contador_loss_consecutivas = 0
                            i=0
                            entrada = float(config['AJUSTES']['valor_entrada'])

                        elif resultado == 0:
                            if i == 0:
                                print('\n>> Resultado: EMPATE \n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2),'\n>> Saldo conta:',cifrao,valorconta)
                            
                            if i >= 1:
                                print('\n>> Resultado: EMPATE no gale',str(i),'\n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2),'\n>> Saldo conta:',cifrao,valorconta)

                            if i+1 <= martingale:
                                gale = float(entrada)                   
                                entrada = round(abs(gale), 2)

                        else:
                            if i == 0:
                                print('\n>> Resultado: LOSS \n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2))
                                #resultado_loss = resultado_loss + 1
                            if i >= 1:
                                print('\n>> Resultado: LOSS no gale',str(i), '\n>> Lucro:', round(resultado,2), '\n>> Par:', ativo, '\n>> Lucro total: ', round(lucro_total,2))
                            resultado_loss = resultado_loss + 1
                            contador_win_consecutivas =0
                            contador_loss_consecutivas += 1
                            if contador_loss_consecutivas > contador_loss_consecutivas_historico:
                                contador_loss_consecutivas_historico = contador_loss_consecutivas

                                """if contador_loss_consecutivas == 3:
                                    print('Entrando em soneca por 30 minutos')
                                    time.sleep(1800) # espera 30 min"""
                                
                            if i+1 <= martingale:
                                
                                gale = float(entrada) * float(fator_mg)                           
                                entrada = round(abs(gale), 2)

                        conta_antiga = conta
                        if conta == 'PRACTICE' and lucro_total <= float('-'+str(abs(stop_loss))): # Se stop loss na conta demo
                            resultado_win = 0
                            resultado_loss = 0
                            lucro_total = 0
                            contador_loss_consecutivas_historico = 0
                            entrada = float(config['AJUSTES']['valor_entrada'])

                        check_stop(entrada)

                        # zerar resultados ao trocar da conta treinamento para real
                        if conta_antiga == 'PRACTICE' and conta == 'REAL':
                            resultado_win = 0
                            resultado_loss = 0
                            lucro_total = 0
                            contador_loss_consecutivas_historico = 0
                            contador_win_consecutivas_historico = 0
                            entrada = float(config['AJUSTES']['valor_entrada'])

                        if conta == 'PRACTICE':
                            if entrada > float(stop_loss) or lucro_total <= float('-'+str(abs(stop_loss))): # Se stop loss
                                resultado_win = 0
                                resultado_loss = 0
                                lucro_total = 0
                                contador_loss_consecutivas_historico = 0
                                contador_win_consecutivas_historico = 0
                                entrada = float(config['AJUSTES']['valor_entrada'])

                        if conta == 'PRACTICE':
                           if lucro_total >= float(abs(stop_win)): #STOP WIN
                                resultado_win = 0
                                resultado_loss = 0
                                lucro_total = 0
                                contador_loss_consecutivas_historico = 0
                                contador_win_consecutivas_historico +=1
                                entrada = float(config['AJUSTES']['valor_entrada'])

                        break


                if resultado > 0:
                    break

            else:
                print('erro na abertura da ordem,', id,ativo)
            break # para o loop

    if soros:
        if lucro_op_atual > 0:
            nivel_soros += 1
            lucro_op_atual = 0
        
        else:
            valor_soros = 0
            nivel_soros = 0
            lucro_op_atual = 0
    return i, entrada, resultado_win, resultado_loss, quantidade_compra, lucro_total,resultado

### Fução que busca hora da corretora ###
def horario():
    x = API.get_server_timestamp()
    now = datetime.fromtimestamp(API.get_server_timestamp())
    
    return now

def medias(velas):
    soma = 0
    for i in velas:
        soma += i['close']
    media = soma / velas_medias

    if media > velas[-1]['close']:
        tendencia = 'put'
    else:
        tendencia = 'call'

    return tendencia
def minutos_restantes_ate_hora_alvo(hora_alvo):
    agora = datetime.now()
    hora_alvo = agora.replace(hour=hora_alvo, minute=0, second=0, microsecond=0)
    
    # Se a hora alvo já passou no mesmo dia, ajusta para o próximo dia
    if agora >= hora_alvo:
        hora_alvo += timedelta(days=1)

    diferenca = hora_alvo - agora
    minutos_restantes = diferenca.total_seconds() / 60
    return int(minutos_restantes)

def pausar_ate_hora(hora_alvo):
    minutos_restantes = minutos_restantes_ate_hora_alvo(hora_alvo)
    print(f"Pausando execução por {minutos_restantes} minutos até as {hora_alvo}...")

    # Aguardar até a hora alvo
    time.sleep(minutos_restantes * 60)

### Função de análise MHI   
def estrategia_mhi():
    global tipo
    # Cria um DataFrame vazio para armazenar os resultados
    VARbd_sqlite = bd_sqlite()
    conn = VARbd_sqlite.conectar_banco()
    VARbd_sqlite.criar_tabela(conn)
    valor_entrada = float(config['AJUSTES']['valor_entrada'])
    compra_realizada = False
    i = 0
    resultado_loss = 0
    resultado_win = 0
    quantidade_compra = 0
    lucro_total = 0
    contador_noticias = 0
    pode_comprar = 'vazio'
    dados_compra = []
    estrategias_realizadas = []

    while True:
        time.sleep(0.1)
        tipo = config['AJUSTES']['tipo']
        
        ### Horario do computador ###
        #minutos = float(datetime.now().strftime('%M.%S')[1:])

        ### horario da iqoption ###
        segundos = int(datetime.fromtimestamp(API.get_server_timestamp()).strftime('%S'))

        entrar = True if segundos >= 59 or segundos <= 0 else False

        print('Aguardando Horário de entrada ', segundos, end='\r')
        
        if entrar:
            # Pausar a execução entre 18 e 19 horas
            #hora_atual = datetime.now().hour
            #if 18 <= hora_atual < 19:
                #pausar_ate_hora(19)
            # teste inicio >>>>>>>>>>>>>>>>>>>

            if contador_noticias <5:
                # Chamada do método para obter notícias
                noticias_investing.obter_noticias(url)
                contador_noticias = 0
            else: 
                contador_noticias = contador_noticias + 1

           

            print('\n>> Iniciando catalogação')
            lista_catalog , linha = catag(API)
            pares_disponiveis = []
            for ativo_atual in lista_catalog:
                 
                pares_disponiveis.append(ativo_atual[0])
                if otc == 'N':
                    if 'OTC' in ativo_atual[0]:
                        print('Encontrado OTC, iniciando soneca por 15 minutos')
                        time.sleep(450) # Espera 15 min
                        break
                # Chamada da função para buscar e avaliar um ativo
                ativo_noticia = ativo_atual[0]
                pode_comprar,df_dados_compra = noticias_investing.buscar_e_avaliar_ativo(ativo_noticia)
                if pode_comprar == 'nao':
                    print(df_dados_compra)
                    print(f'Pares disponiveis: {pares_disponiveis}')
                    continue

                #ativo = lista_catalog[0][0]
                #assertividade = lista_catalog[0][linha]
                ativo = ativo_atual[0]
                assertividade = ativo_atual[linha]
                #print('\n>> Melhor par: ', ativo, ' | Assertividade: ', assertividade)

                if tipo == 'automatico':
                    binary, turbo, digital = payout(ativo)
                    print(binary, turbo, digital )
                    if digital > turbo:
                        print( '\n>> Suas entradas serão realizadas nas digitais')
                        tipo = 'digital'
                    elif turbo > digital:
                        print( '\n>> Suas entradas serão realizadas nas binárias')
                        tipo = 'binary'
                    else:
                        if binary > 80:
                            tipo = 'binary'
                            print('\n>> Os valores de payout estão iguais, escolhido binaria')
                        elif digital > 80:
                            tipo = 'digital'
                            print('\n>> Os valores de payout estão iguais, escolhido digital')
                        else:
                            print('\n>> Os valores de payout estão iguais, não foi escolhido nenhum')
                            break
                        # teste fim >>>>>>>>>>>>>>>>>>>
                    print('\n>> Iniciando análise da estratégia MHI')

                    direcao = False

                    timeframe = 60
                    qnt_velas = 7 # !!! estava 3
                    

                    if analise_medias == 'S':
                        velas = API.get_candles(ativo, timeframe, velas_medias, time.time())
                        #velas = API.get_candles(ativo, timeframe, qnt_velas, time.time())
                        tendencia = medias(velas)

                    else:
                        velas = API.get_candles(ativo, timeframe, qnt_velas, time.time())

                    #velas = ['Verde' if vela['open'] < vela['close'] else 'Vermelha' if vela['open'] > vela['close'] else 'Doji' for vela in velas]

                    # Estratégias
                    estrategias = [
                        ('mhi um minoria', velas[-3:], 'put' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'call', None),
                        ('mhi um maioria', velas[-3:], 'call' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'put', None),
                        ('milhao maioria', velas[-5:], 'call' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'put', None),
                        ('milhao minoria', velas[-5:], 'put' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'call', None),
                        ('mhi dois minoria', velas[-5:-2], 'put' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'call', None),
                        ('mhi dois maioria', velas[-5:-2], 'put' if velas.count('Verde') > velas.count('Vermelha') and 'Doji' not in velas else 'put', None)
                    ]
    # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< !!!! ATIVAR QUANDO FOR REAL !!!! >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
                    for i, (estrategia_nome, velas_analisadas, direcao_esperada, _) in enumerate(estrategias):
                    #for i, (estrategia_nome, velas_analisadas, direcao_esperada, chance) in enumerate(estrategias): # retirar pós teste

                        chance = VARbd_sqlite.listar_estrategias_por_vitoria()

                        estrategias[i] = (estrategia_nome, velas_analisadas, direcao_esperada, chance)

                    # Organize as estratégias com base na chance em ordem decrescente ATIVAR
                    estrategias_ordenadas = sorted(estrategias, key=lambda x: x[3][1], reverse=True)

                    """if 'OTC' in ativo_atual[0]:
                        # Organize as estratégias com base na chance em ordem crescente
                        estrategias_ordenadas = sorted(estrategias, key=lambda x: x[3][1])
                    else:
                        # Organize as estratégias com base na chance em ordem decrescente
                        estrategias_ordenadas = sorted(estrategias, key=lambda x: x[3][1], reverse=True)"""


                    # Agora estrategias_ordenadas contém as estratégias organizadas pela ordem escolhida acima
                    for estrategia in estrategias_ordenadas:
                    
                        estrategia_nome, velas_analisadas, direcao_esperada, chance = estrategia
                        porcentagem_vitoria = next((item[1] for item in chance[0] if item[0] == estrategia_nome), 0)

                        if resultado_loss > 0:
                            #print(f"Estratégia: {estrategia_nome}, Direção Esperada: {direcao_esperada}, Chance: {porcentagem_vitoria:.2f}%. Não realizado entrada!!!")
                            #continue
                            if estrategia_nome in estrategias_realizadas:
                                continue
                            else:
                                estrategias_realizadas.append(estrategia_nome)

                            if len(estrategias_realizadas) == 6:
                                estrategias_realizadas = []
                       
                        if analise_medias == 'S':
                            if direcao_esperada != tendencia:
                                print(f'Estratégia {estrategia_nome}: Entrada abortada - Contra Tendência. Ativo: {ativo_atual[0]} Direção esperada: {direcao_esperada} e Tendencia: {tendencia}')
                                continue 

                        if direcao_esperada == 'put' or direcao_esperada == 'call':


                            print(f"Estratégia: {estrategia_nome}, Direção Esperada: {direcao_esperada}, Chance: {porcentagem_vitoria:.2f}%")
                            i, entrada, resultado_win, resultado_loss, quantidade_compra, lucro_total,resultado = compra(ativo, valor_entrada, direcao_esperada, 1, tipo, i, resultado_win, resultado_loss, quantidade_compra)
                            # guardar em um banco sqlite o ativo_atual[0], resultado e estrategia_nome
                            if resultado > 0:
                                resultado_extenso = 'Win'
                            else:
                                resultado_extenso = 'Loss'
                            conn = VARbd_sqlite.conectar_banco()
                            VARbd_sqlite.inserir_dados(conn,estrategia_nome,ativo_atual[0],resultado_extenso)


                            valor_entrada = entrada
                            if direcao_esperada == 'put' or direcao_esperada == 'call':
                                break  # Sai do loop se a direção for 'put' ou 'call'

                        else:
                            print(f'Estratégia {estrategia_nome}: Entrada abortada - Foi encontrado um doji na análise.')
                            time.sleep(2)
                    time.sleep(1)
                    valorconta = float(API.get_balance())

                    print('\n######################################################################')
                    print('\nOlá, ',nome, '\nSeja bem vindo ao Robô do Girard.')
                    print('>>> Seu Saldo na conta ',conta, 'é de', cifrao,valorconta)
                    print('Seu valor de entrada é de ',cifrao,valor_entrada)
                    print('Stop win:',cifrao,stop_win)
                    print('Stop loss:',cifrao,'-',stop_loss)
                    print('Resultado WIN:',resultado_win)
                    print('Resultado LOSS:',resultado_loss)
                    if resultado_win > 0:
                        caluclo_assertividade = (resultado_win/quantidade_compra)*100
                        print(f'>>> A assertividade é de {caluclo_assertividade:.2f}%')
                    else:
                        print(f'>>> A assertividade é de 0%')
                    print(f'Lucro total: {lucro_total}')
                    print(f'>>>>>> Perdas consecutivas: {contador_loss_consecutivas_historico}')
                    print('\n######################################################################\n')
                    if conta == 'REAL':
                        if contador_loss_consecutivas_historico >= 3:
                            print('Alcançado o limite diário de perdas, tente amanhã!')
                            sys.exit()


### DEFININCãO INPUTS NO INICIO DO ROBÔ ###

perfil = json.loads(json.dumps(API.get_profile_ansyc()))
cifrao = str(perfil['currency_char'])
nome = str(perfil['name'])
valorconta = float(API.get_balance())

print('\n######################################################################')
print('\nOlá, ',nome, '\nSeja bem vindo ao Robô do Girard.')
print('\nSeu Saldo na conta ',conta, 'é de', cifrao,valorconta)
print('\nSeu valor de entrada é de ',cifrao,valor_entrada)
print('\nStop win:',cifrao,stop_win)
print('\nStop loss:',cifrao,'-',stop_loss)
print('\n######################################################################\n\n')

### chamada da estrategia mhi ###
estrategia_mhi()